import io
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font


def generate_excel_bytes(data_list):
    """Build formatted Excel file in memory for download."""
    if not data_list:
        return None, None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"google_maps_data_{timestamp}.xlsx"

    df = pd.DataFrame(data_list)

    expected_cols = ['name', 'phone', 'email', 'facebook', 'instagram', 'twitter', 'linkedin', 'youtube', 'tiktok', 'website', 'address', 'rating', 'reviews', 'category', 'hours', 'files_url']
    for col in expected_cols:
        if col not in df.columns:
            df[col] = "N/A"

    export_df = df[expected_cols].copy()
    export_df.columns = ['Name', 'Phone', 'Email', 'Facebook', 'Instagram', 'Twitter', 'LinkedIn', 'YouTube', 'TikTok', 'Website', 'Address', 'Rating', 'Reviews', 'Category', 'Hours', 'Map URL']

    buffer = io.BytesIO()
    export_df.to_excel(buffer, index=False, sheet_name='Business Data')
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb['Business Data']

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    ws_summary = wb.create_sheet('Summary')
    ws_summary['A1'] = 'Total Records Scraped'
    ws_summary['B1'] = len(data_list)
    ws_summary['A2'] = 'Scrape Date'
    ws_summary['B2'] = timestamp

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), filename
